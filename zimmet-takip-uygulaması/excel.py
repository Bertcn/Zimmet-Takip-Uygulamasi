import openpyxl
import sys  
from PyQt5.QtWidgets import QApplication, QMainWindow
from panel import Ui_MainWindow
from PyQt5.QtWidgets import QMessageBox
from PyQt5 import QtCore
from openpyxl.styles import Border, Side
from datetime import datetime

app = QApplication(sys.argv)

pencere = QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(pencere)
pencere.show()

# Kaynak Excel dosyasını aç
workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook['SUBE_VE_GENEL']

target_workbook = openpyxl.load_workbook('example.xlsx')  # Hedef dosyanın adını belirtin
target_sheet =workbook['GELECEK_ENVANTER']

# Başlangıç ID değer
baslangic_id = sheet.max_row + 1

# Tablodaki mevcut çalışan numaralarını saklayın
calisan_nolar = {row[0] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)}
pcbarkodlar = {row[7] for row in sheet.iter_rows(min_row=2, values_only=True)}


for i in range(1, 1001):
    ui.calisannocombo.addItem(str(i))  # Sayıları QComboBox'a ekleyin
sube_listesi = [
    "ADANA SEYHAN ŞUBE",
    "AFYONKARAHİSAR ŞUBE",
    "ANKARA ETİMESGUT ŞUBESİ",
    "ANKARA KEÇİÖREN ŞUBE",
    "ANKARA MAMAK ŞUBE",
    "ANKARA SIHHİYE ŞUBE",
    "ANKARA SİNCAN ŞUBE",
    "ANTALYA KEPEZ ŞUBE",
    "ANTALYA MURATPAŞA ŞUBE",
    "AVCILAR ŞUBE",
    "AYDIN ŞUBE",
    "BAĞCILAR ŞUBE",
    "BALIKESİR ŞUBE",
    "BATMAN ŞUBE",
    "BURSA ANKARA YOLU CADDESİ ŞUBESİ",
    "BURSA İNEGÖL ŞUBE",
    "BURSA OSMANGAZİ ŞUBE",
    "ÇANAKKALE ŞUBE",
    "DENİZLİ ŞUBE",
    "ELAZIĞ ŞUBE",
    "ERZURUM ŞUBE",
    "ESENLER ŞUBE",
    "ESENYURT ŞUBE",
    "ESKİŞEHİR ŞUBE",
    "FATİH FEVZİPAŞA ŞUBE",
    "GAZİANTEP KARAGÖZ ŞUBE",
    "GAZİOSMANPAŞA ŞUBE",
    "GEBZE ŞUBE",
    "GENEL MÜDÜRLÜK",
    "ISPARTA ŞUBE",
    "İZMİR BORNOVA ŞUBE",
    "İZMİR ÇANKAYA ŞUBE",
    "İZMİR KARABAĞLAR",
    "KARTAL ŞUBE",
    "KAYSERİ MELİKGAZİ ŞUBE",
    "KOCAELİ İZMİT FEVZİYE ŞUBE",
    "KONYA SELÇUKLU ŞUBE",
    "KONYA ŞUBE",
    "KÜTAHYA ŞUBE",
    "LEVENT ŞUBE",
    "MANİSA YUNUS EMRE ŞUBE",
    "MARDİN ŞUBE",
    "MERSİN AKDENİZ ŞUBE",
    "MERSİN TARSUS ŞUBE",
    "ORDU ŞUBE",
    "OSMANİYE ŞUBE",
    "PENDİK ŞUBE",
    "SAKARYA ŞUBE",
    "SAMSUN ŞUBE",
    "SANCAKTEPE ŞUBE",
    "SİVAS ŞUBE",
    "SULTANBEYLİ ŞUBE",
    "SULTANGAZİ ŞUBE",
    "ŞANLIURFA ŞUBE",
    "ŞİRİNEVLER ŞUBE",
    "TATVAN ŞUBE",
    "TEKİRDAĞ ÇORLU ŞUBE",
    "TOKAT ŞUBE",
    "TRABZON ŞUBE",
    "UŞAK ŞUBE",
    "ÜMRANİYE ŞUBE",
    "ÜSKÜDAR ŞUBE",
    "VAN ŞUBE"
]
for sube in sube_listesi:
    ui.comboBox.addItem(sube)

sube_baslangic_sira = {
    "ADANA SEYHAN ŞUBE": 4,
    "AFYONKARAHİSAR ŞUBE": 22,
    "ANKARA ETİMESGUT ŞUBESİ": 40,
    "ANKARA KEÇİÖREN ŞUBE": 58,
    "ANKARA MAMAK ŞUBE": 76,
    "ANKARA SIHHİYE ŞUBE": 94,
    "ANKARA SİNCAN ŞUBE": 112,
    "ANTALYA KEPEZ ŞUBE": 130,
    "ANTALYA MURATPAŞA ŞUBE": 148,
    "AVCILAR ŞUBE": 166,
    "AYDIN ŞUBE": 184,
    "BAĞCILAR ŞUBE": 202,
    "BALIKESİR ŞUBE": 220,
    "BATMAN ŞUBE": 238,
    "BURSA ANKARA YOLU CADDESİ ŞUBESİ": 256,
    "BURSA İNEGÖL ŞUBE": 274,
    "BURSA OSMANGAZİ ŞUBE": 292,
    "ÇANAKKALE ŞUBE": 310,
    "DENİZLİ ŞUBE": 328,
    "ELAZIĞ ŞUBE": 346,
    "ERZURUM ŞUBE": 364,
    "ESENLER ŞUBE": 382,
    "ESENYURT ŞUBE": 400,
    "ESKİŞEHİR ŞUBE": 418,
    "FATİH FEVZİPAŞA ŞUBE": 436,
    "GAZİANTEP KARAGÖZ ŞUBE": 454,
    "GAZİOSMANPAŞA ŞUBE": 472,
    "GEBZE ŞUBE": 490,
    "ISPARTA ŞUBE": 508,
    "İZMİR BORNOVA ŞUBE": 526,
    "İZMİR ÇANKAYA ŞUBE": 544,
    "İZMİR KARABAĞLAR": 562,
    "KARTAL ŞUBE": 580,
    "KAYSERİ MELİKGAZİ ŞUBE": 598,
    "KOCAELİ İZMİT FEVZİYE ŞUBE": 616,
    "KONYA SELÇUKLU ŞUBE": 634,
    "KONYA ŞUBE": 652,
    "KÜTAHYA ŞUBE": 670,
    "LEVENT ŞUBE": 688,
    "MANİSA YUNUS EMRE ŞUBE": 706,
    "MARDİN ŞUBE": 724,
    "MERSİN AKDENİZ ŞUBE": 742,
    "MERSİN TARSUS ŞUBE": 760,
    "ORDU ŞUBE": 778,
    "OSMANİYE ŞUBE": 796,
    "PENDİK ŞUBE": 814,
    "SAKARYA ŞUBE": 832,
    "SAMSUN ŞUBE": 850,
    "SANCAKTEPE ŞUBE": 868,
    "SİVAS ŞUBE": 886,
    "SULTANBEYLİ ŞUBE": 904,
    "SULTANGAZİ ŞUBE": 922,
    "ŞANLIURFA ŞUBE": 940,
    "ŞİRİNEVLER ŞUBE": 958,
    "TATVAN ŞUBE": 976,
    "TEKİRDAĞ ÇORLU ŞUBE": 994,
    "TOKAT ŞUBE": 1012,
    "TRABZON ŞUBE": 1030,
    "UŞAK ŞUBE": 1048,
    "ÜMRANİYE ŞUBE": 1066,
    "ÜSKÜDAR ŞUBE": 1084,
    "VAN ŞUBE": 1102,
    "GENEL MÜDÜRLÜK": 1120
}

for sube_adi in sube_baslangic_sira:
        sube_baslangic_sira[sube_adi] += 0
    

def find_empty_row(sheet, start_row):
    for row in range(start_row, sheet.max_row + 2):
        if all(cell.value is None for cell in sheet[row]):
            return row
    return sheet.max_row + 1

tum_veriler=[]

def kayitekle():
    global baslangic_id
    
    No = int(ui.calisannocombo.currentText())
    sube = ui.comboBox.currentText()
    Pcbarkod =ui.pcbarkod.text()

    if not Pcbarkod.strip():
        error_message = "PCBARKOD kısmı boş bırakıldı."
        QMessageBox.critical(ui.centralwidget, "Hata", error_message)
        return
    
    try:
        Pcbarkod = int(Pcbarkod)
    except ValueError:
        error_message = "PCBARKOD kısmına tam sayı girilmedi."
        QMessageBox.critical(ui.centralwidget, "Hata", error_message)
        return
    
    
    if No in calisan_nolar:
        error_message = "Bu çalışan numarası zaten mevcut. Kayıt eklenmedi."
        QMessageBox.critical(ui.centralwidget, "Hata", error_message)
        return  # Kayıt eklemeyi durdur
    
    
    if Pcbarkod in pcbarkodlar:
        error_message = "Bu Barkod numarası zaten mevcut. Kayıt eklenmedi."
        QMessageBox.critical(ui.centralwidget, "Hata", error_message)
        return  # Kayıt eklemeyi durdur
    
    


    for sube_adi in sube_baslangic_sira:
        sube_baslangic_sira[sube_adi] += 0

    
    # example dosyasında boş satır arayın
    bos_satir = find_empty_row(sheet, sube_baslangic_sira[sube])
    
    Ad = ui.adsoyad.text()
    cihazdurumu = ui.cihazdurumu.text()
    Pcmarka = ui.pcmarka.text()
    pcmodel = ui.pcmodel.text()
    Pcimei = ui.lineEdit_4.text()
    
    islemci = ui.islemci.currentText()
    ssd = ui.SSD.text()
    ram = ui.ram.currentText()
    Telmarka = ui.telmarka.text()
    Telmodel=ui.telmodel.text()
    Telimei = ui.telimei.text()
    Telbarkod = ui.lineEdit.text()
    telserino = ui.telserino.text()
    Tarih = ui.dateEdit.text()
    İslemtarihi=datetime.now()

    liste = [No, Ad, sube, cihazdurumu, Pcmarka, pcmodel, Pcimei, Pcbarkod, islemci, ssd, ram, Telmarka,Telmodel,telserino,Telimei, Telbarkod, Tarih,İslemtarihi]
    tum_veriler.append(liste)
    
    # example dosyasına verileri ekle
    for col, value in enumerate(liste, start=1):
        sheet.cell(row=bos_satir, column=col, value=value)
        if col != 2:  # Ad-Soyad hücresi hariç diğer hücrelerin çerçevesini ayarla
            border_style = Border(left=Side(border_style="thin"), 
                                  right=Side(border_style="thin"), 
                                  top=Side(border_style="thin"), 
                                  bottom=Side(border_style="thin"))
            sheet.cell(row=bos_satir, column=col).border = border_style
    
    # Değişiklikleri kaydet
    workbook.save('example.xlsx')
    
    # Sonraki kayıt için yeni bir ID değeri
    calisan_nolar.add(No)  # Yeni çalışan numarasını sete ekleyin
    pcbarkodlar.add(Pcbarkod)

    success_message = "Kayıt başarıyla eklendi."
    QMessageBox.information(ui.centralwidget, "Başarılı", success_message)

def calisansil():
    sil_calisan = int(ui.calisannocombo.currentText())  # Silinecek çalışan numarası

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == sil_calisan:
            if row[2] == "GENEL MÜDÜRLÜK":
                target_empty_row = find_empty_row(target_sheet, start_row=1)
                for col, value in enumerate(row, start=1):
                    target_sheet.cell(row=target_empty_row, column=col, value=value)
                target_workbook.save('target_example.xlsx')
                for cell in sheet[row_number]:  # Satırı komple boşalt
                    cell.value = None
                success_message = "Çalışan bilgileri güncellendi ve taşındı."
                QMessageBox.information(ui.centralwidget, "Başarılı", success_message)
            else:
                ad_soyad_cell = sheet.cell(row=row_number, column=2)
                ad_soyad_cell.value = "BOŞ"
                ad_soyad_cell.font = openpyxl.styles.Font(color="FF0000")
                
                target_empty_row = find_empty_row(target_sheet, start_row=1)
                for col, value in enumerate(row, start=1):
                    target_sheet.cell(row=target_empty_row, column=col, value=value)
                target_workbook.save('GELECEK_ENVANTER')
                success_message = "Çalışan bilgileri güncellendi ve taşındı."
                QMessageBox.information(ui.centralwidget, "Başarılı", success_message)
            break
    else:
        print("Böyle bir kayıt yok")

    workbook.save('example.xlsx')


def pcbarkodilesil():
    barkod_input = ui.pcbarkod.text() # Silinecek cihazın barkodu
    try:
        sil_barkod=int(barkod_input)
    except ValueError:
        QMessageBox.warning(ui.centralwidget,"HATA","BARKOD NUMARASI SAYI OLMALIDIR")
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[7] == sil_barkod:
            if row[2] == "GENEL MÜDÜRLÜK":
                target_empty_row = find_empty_row(target_sheet, start_row=1)
                for col, value in enumerate(row, start=1):
                    target_sheet.cell(row=target_empty_row, column=col, value=value)
                target_workbook.save('target_example.xlsx')
                for cell in sheet[row_number]:
                    cell.value = None
                success_message = "Cihaz bilgileri güncellendi ve taşındı."
                QMessageBox.information(ui.centralwidget, "Başarılı", success_message)
            else:
                ad_soyad_cell = sheet.cell(row=row_number, column=2)
                if ad_soyad_cell.value != "BOŞ":  # Ad soyad doluysa silme işlemi yap
                    ad_soyad_cell.value = "BOŞ"
                    ad_soyad_cell.font = openpyxl.styles.Font(color="FF0000")
                    target_empty_row = find_empty_row(target_sheet, start_row=1)
                    for col, value in enumerate(row, start=1):
                        target_sheet.cell(row=target_empty_row, column=col, value=value)
                    target_workbook.save('target_example.xlsx')
                    success_message = "Cihaz bilgileri güncellendi ve taşındı."
                    QMessageBox.information(ui.centralwidget, "Başarılı", success_message)
                else:
                    QMessageBox.warning(ui.centralwidget, "Uyarı", f"{sil_barkod} barkodlu cihazın ad soyad bilgisi zaten BOŞ.")
            break
    else:
                    QMessageBox.warning(ui.centralwidget, "Uyarı", f"{sil_barkod} barkodlu cihaz yok.")

    workbook.save('example.xlsx')


def tayinet():
    calisan_no = int(ui.calisannocombo.currentText())  # Tayin edilecek çalışan numarası
    yeni_sube = ui.comboBox.currentText()  # Hedef şube adı

    # Çalışanı bul ve verilerini al
    calisan_index = None
    calisan_data = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == calisan_no:
            calisan_index = row_number
            calisan_data = row
            break

    if calisan_index is None:
        QMessageBox.warning(ui.centralwidget, "Uyarı", f"{calisan_no} numaralı çalışan kaydı bulunamadı.")
        return

    # Eğer çalışan zaten hedef şubede ise hata mesajı ver
    if calisan_data[2] == yeni_sube:
        QMessageBox.warning(ui.centralwidget, "Uyarı", f"{calisan_no} numaralı çalışan zaten {yeni_sube} şubesinde.")
        return

    # Hedef şubenin altındaki boş satırı bul
    target_row = find_empty_row(sheet, sube_baslangic_sira[yeni_sube])
    for col, value in enumerate(calisan_data, start=1):
        sheet.cell(row=target_row, column=col, value=value)

    # Çalışanın eski satırını temizle
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=calisan_index, column=col).value = None

    # Çalışanın şube adını güncelle
    sheet.cell(row=target_row, column=3, value=yeni_sube)

    success_message = f"Çalışan {calisan_no} numaralı çalışan {yeni_sube} şubesine tayin edildi."
    QMessageBox.information(ui.centralwidget, "Başarılı", success_message)
    workbook.save('example.xlsx')  # Değişiklikleri kaydet

def subedeyeni():
    pc_barkod = int(ui.pcbarkod.text())  # Hedef PC barkodu
    yeni_ad_soyad = ui.adsoyad.text()  # Yeni ad soyad
    yeni_calisan_no = int(ui.calisannocombo.currentText())  # Yeni çalışan numarası

    # Hedef PC barkodunun olduğu satırı bul
    target_row = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[7] == pc_barkod:  # Eğer hedef PC barkodunu bulursak
            target_row = row_number
            break

    if target_row is None:
        QMessageBox.warning(ui.centralwidget, "Uyarı", f"{pc_barkod} barkodlu cihaz kaydı bulunamadı.")
        return
    
    # Eğer yeni_calisan_no zaten tabloda mevcutsa hata ver ve yeni bir değer girmesini iste
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == yeni_calisan_no:
            QMessageBox.warning(ui.centralwidget, "Uyarı", f"{yeni_calisan_no} çalışan numarası zaten tabloda mevcut. Lütfen farklı bir numara girin.")
            return

    # Eğer ad soyad kısmı "BOŞ" ise yeni verileri ekleyin
    if sheet.cell(row=target_row, column=2).value == "BOŞ":
        sheet.cell(row=target_row, column=1, value=yeni_calisan_no)
        sheet.cell(row=target_row, column=2, value=yeni_ad_soyad)
        success_message = f"{pc_barkod} barkodlu cihazın ad soyad bilgileri güncellendi."
        QMessageBox.information(ui.centralwidget, "Başarılı", success_message)
    else:
        QMessageBox.warning(ui.centralwidget, "Uyarı", f"{pc_barkod} barkodlu cihazın ad soyad bilgisi dolu.")
    
    workbook.save('example.xlsx')  # Değişiklikleri kaydet




ui.imeisilbutton.clicked.connect(pcbarkodilesil)
ui.buton.clicked.connect(kayitekle)
ui.calisannosil.clicked.connect(calisansil)
ui.pushButton.clicked.connect(tayinet)
ui.subedeyeni.clicked.connect(subedeyeni)

if __name__ == "__main__":
    sys.exit(app.exec_())
#burası çalışıyorrrrr......

#ANA DOSYA BU 